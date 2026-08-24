import re
import numpy as np
import pandas as pd
import openpyxl
from app.core.config import MASTER_COLUMNS_PREMI, MASTER_COLUMNS_CLAIM
from app.utils.helpers import read_excel_dynamic_header, to_snake_case, validate_dates

class TripakartaETL:
    """Modul khusus pemrosesan data Tripakarta (Premi & Klaim)"""

    @staticmethod
    def process_premi(file_path: str, target_sheet: str, periode_lengkap: str, override_cob: str = None) -> pd.DataFrame:
        master_cols = MASTER_COLUMNS_PREMI["tripakarta"]
        
        # 1. Ekstrak Treaty Name dari Header Atas
        df_top = pd.read_excel(file_path, sheet_name=target_sheet, nrows=6, header=None)
        treaty_raw = "Fire Quota Share Surplus"
        for i in range(len(df_top)):
            for j in range(len(df_top.columns)):
                cell_val = str(df_top.iloc[i, j]).strip()
                if "Treaty Name" in cell_val:
                    for k in range(j+1, len(df_top.columns)):
                        val_kanan = str(df_top.iloc[i, k]).strip()
                        if val_kanan and val_kanan.lower() != 'nan':
                            treaty_raw = val_kanan
                            break
                    break

        if "Quota Share Surplus" in treaty_raw:
            jenis = treaty_raw.split(" ")[0]
            val_qs = f"{jenis} Quota Share"
            val_surplus = f"{jenis} Surplus"
        else:
            val_qs = treaty_raw
            val_surplus = treaty_raw

        # 2. Baca Data
        df = read_excel_dynamic_header(file_path, target_sheet)

        # 3. Mapping Kolom Berdasarkan Nama Header Excel Tripakarta
        rename_dict = {
            'No': 'no',
            'COB': 'cob',
            'REINSURED': 'reinsured',
            'POLICY NUMBER': 'policy_number',
            'INSURED NAME': 'insured_name',
            'UW YEAR': 'uw_year',
            'CURRENCY': 'currency',
            
            # Breakdown of SI (Merged Header)
            'BREAKDOWN OF SI': 'breakdown_of_si_md_building',
            'Unnamed: 9': 'mb',
            'Unnamed: 10': 'stock',
            'Unnamed: 11': 'tpl',
            'Unnamed: 12': 'bi',
            'Unnamed: 13': 'other',
            
            '100% TSI': 'tsi_100',
            'BASIS OF INDEMNITY': 'basis_of_indemnity',
            'OCCUPATION CODE': 'occupation_code',
            'OCCUPATION': 'occupation',
            'LOCATION': 'location',
            
            # Period of Insurance (Merged Header: Start & End)
            'PERIOD OF INSURANCE': 'period_of_insurance_start',
            'Unnamed: 20': 'period_of_insurance_end',
            
            'SOURCE (DIRECT/COINS/INWARD FAC.)': 'source_direct_coins_inward_fac',
            "CEDANT'S SHARE": 'cedant_s_share',
            
            # Spreading of Risk (Merged Header: OR, QS, Surplus, Others)
            'SPREADING OF RISK': 'spreading_of_risk_or',
            'Unnamed: 24': 'qs',
            'Unnamed: 25': 'surplus',
            'Unnamed: 26': 'others',
            
            '100% Premium': 'premium_100',
            'Premium Rate ': 'premium_rate',
            'Premi QS': 'premi_qs',
            'Comm QS': 'comm_qs',
            'Premi SPL': 'premi_spl',
            'Comm SPL': 'comm_spl',
            '% Marsh Re': 'marsh_re',
            'Premium QS\n(Marsh Re Share)': 'premium_qs_marsh_re_share',
            'Premium SPL\n(Marsh Re Share)': 'premium_spl_marsh_re_share',
            'NOTE': 'note',
            'REMARKS': 'remarks'
        }

        df.rename(columns=rename_dict, inplace=True)
        if 'Unnamed: 0' in df.columns:
            df.drop(columns=['Unnamed: 0'], inplace=True)

        # 4. Buang Baris Sub-Header yang bukan data (misal teks 'START', 'END', 'MD/BUILDING')
        if 'period_of_insurance_start' in df.columns:
            mask_sub = df['period_of_insurance_start'].astype(str).str.upper().str.contains(r'START|INCEPTION|PERIODE', na=False)
            df = df[~mask_sub]

        # 5. Parsing dan Format Kolom Tanggal ke Timestamp
        for dcol in ['period_of_insurance_start', 'period_of_insurance_end']:
            if dcol in df.columns:
                def clean_date_entry(val):
                    if pd.isna(val) or val is None:
                        return None
                    if isinstance(val, (pd.Timestamp, np.datetime64)):
                        return pd.to_datetime(val).strftime('%Y-%m-%d %H:%M:%S')

                    val_str = str(val).strip()
                    if val_str.lower() in ['', 'nan', 'nat', 'none', 'null', '-', '0']:
                        return None

                    # Serial Excel
                    try:
                        num_v = float(val_str)
                        if 30000 <= num_v <= 65000:
                            dt = pd.to_datetime('1899-12-30') + pd.to_timedelta(num_v, unit='D')
                            return dt.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        pass

                    # String Tanggal Biasa
                    try:
                        dt_obj = pd.to_datetime(val_str, errors='coerce')
                        if pd.isna(dt_obj) or dt_obj.year < 1900 or dt_obj.year > 2500:
                            return None
                        return dt_obj.strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        return None

                df[dcol] = df[dcol].apply(clean_date_entry)

        # Salin ke alias alternatif jika master column memakai format pendek
        if 'period_of_insurance_start' in df.columns:
            df['period_of_start'] = df['period_of_insurance_start']
            df['start_date'] = df['period_of_insurance_start']
            df['start'] = df['period_of_insurance_start']

        if 'period_of_insurance_end' in df.columns:
            df['period_of_end'] = df['period_of_insurance_end']
            df['end_date'] = df['period_of_insurance_end']
            df['end'] = df['period_of_insurance_end']

        # 6. Metadata & ID Sanitization
        if 'no' in df.columns:
            df['no'] = df['no'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            df['no'] = df['no'].replace(['nan', 'None', 'NaN', 'NaT', ''], np.nan)
        if 'policy_number' in df.columns:
            df['policy_number'] = df['policy_number'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            df['policy_number'] = df['policy_number'].replace(['nan', 'None', 'NaN', 'NaT', ''], np.nan)
        if 'uw_year' in df.columns:
            df['uw_year'] = df['uw_year'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            df['uw_year'] = df['uw_year'].replace(['nan', 'None', 'NaN', 'NaT', ''], np.nan)

        df['period'] = periode_lengkap
        df['treaty_name_qs'] = val_qs
        df['treaty_name_surplus'] = val_surplus

        if override_cob and override_cob.strip() and override_cob.strip().lower() != "string":
            if 'cob' in df.columns:
                df['cob'] = override_cob.strip().upper()

        # 7. Sinkronisasi dengan Master Columns Tripakarta
        for col in master_cols:
            if col not in df.columns:
                df[col] = np.nan

        df_clean = df[master_cols].copy()

        # 8. Filter Baris Footer / Total
        if "policy_number" in df_clean.columns:
            df_clean = df_clean.dropna(subset=["policy_number"], how="all")
            trash_regex = r'^\s*(TOTAL|JUMLAH|GRAND\s*TOTAL|SUBTOTAL|REKAP|SUMMARY|0|NAN|NONE)\s*$'
            pol_str = df_clean["policy_number"].astype(str).str.upper().str.strip()
            df_clean = df_clean[~pol_str.str.match(trash_regex, na=False)]

        # 9. Sanitasi Kolom Numerik
        num_cols = [
            "breakdown_of_si_md_building", "mb", "stock", "tpl", "bi", "other",
            "tsi_100", "100_tsi", "cedant_s_share", "spreading_of_risk_or", "qs", "surplus",
            "others", "premium_100", "100_premium", "premium_rate", "premi_qs", "comm_qs",
            "premi_spl", "comm_spl", "marsh_re", "premium_qs_marsh_re_share",
            "premium_spl_marsh_re_share"
        ]
        for col in num_cols:
            if col in df_clean.columns:
                s = df_clean[col].astype(str).str.strip()
                s = s.replace({'-': np.nan, 'NIL': np.nan, 'nil': np.nan, 'None': np.nan, 'nan': np.nan, 'NaN': np.nan, '': np.nan})
                s = s.str.replace(',', '', regex=False)
                df_clean[col] = pd.to_numeric(s, errors='coerce')

        return df_clean.reset_index(drop=True)

    @staticmethod
    def process_claim(file_path: str, target_sheet: str, periode_lengkap: str, override_cob: str = None) -> pd.DataFrame:
        master_cols = MASTER_COLUMNS_CLAIM["tripakarta"]

        # 1. Buka Excel & Unmerge Semua Merged Cell
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_dict = {s.lower().strip(): s for s in wb.sheetnames}
        actual_sheet = sheet_dict.get(target_sheet.lower().strip(), target_sheet)
        ws = wb[actual_sheet]

        for merge in list(ws.merged_cells):
            min_col, min_row, max_col, max_row = merge.min_col, merge.min_row, merge.max_col, merge.max_row
            top_val = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = top_val

        data_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # 2. Cari Posisi Baris Header Utama
        header_idx = -1
        for i, row in enumerate(data_rows):
            row_str = " ".join([str(x).upper() for x in row if x is not None])
            if any(k in row_str for k in ["REGISTER", "POLICY", "NO POLIS", "NO KLAIM"]):
                header_idx = i
                break

        if header_idx == -1:
            raise ValueError(f"Baris header tidak ditemukan di sheet {target_sheet}")

        # 3. Gabungkan Header Baris 1 & 2
        row_top = data_rows[header_idx]
        row_sub = data_rows[header_idx + 1] if header_idx + 1 < len(data_rows) else []

        raw_headers = []
        for top, sub in zip(row_top, row_sub):
            top_s = re.sub(r'[\r\n\t]+', ' ', str(top)).strip() if top is not None else ""
            sub_s = re.sub(r'[\r\n\t]+', ' ', str(sub)).strip() if sub is not None else ""
            if top_s and sub_s and top_s != sub_s:
                raw_headers.append(f"{top_s} {sub_s}")
            elif top_s:
                raw_headers.append(top_s)
            elif sub_s:
                raw_headers.append(sub_s)
            else:
                raw_headers.append("")

        df = pd.DataFrame(data_rows[header_idx + 2:], columns=raw_headers)
        df = df.loc[:, [c for c in df.columns if c != ""]].copy()

        # 4. Kamus Sinonim Dinamis Presisi
        SYNONYMS = {
            "no": ["no", "nomor", "no_urut"],
            "register_no": ["register_no", "claim_no", "claim_ref_no", "no_klaim", "no_register", "reg_no"],
            "policy_number": ["policy_number", "policy_no", "no_polis", "nomor_polis", "policyno", "nopol"],
            "insured_name": ["insured_name", "nama_tertanggung", "name_of_insured", "insured"],
            "cob": ["cob", "type_of_cover", "class_of_business", "jenis_asuransi", "cover", "line_of_business"],
            "uw_year": ["uw_year", "underwriting_year", "tahun_uw", "uw_yr"],
            "period_of_insurance_start": [
                "period_of_insurance_start", "period_of_asuransi_start", "period_asuransi_start",
                "periode_asuransi_mulai", "period_start", "insurance_start", "inception", "sdate",
                "period_from", "periode_mulai", "tgl_mulai_asuransi"
            ],
            "period_of_insurance_end": [
                "period_of_insurance_end", "period_of_asuransi_end", "period_asuransi_end",
                "periode_asuransi_akhir", "period_end", "insurance_end", "expiry", "edate",
                "period_to", "periode_selesai", "tgl_akhir_asuransi"
            ],
            "occupation_code": ["occupation_code", "kode_okupasi", "occ_code", "kode_usaha"],
            "occupation": ["occupation", "okupasi", "jenis_usaha", "bidang_usaha"],
            "dol": ["dol", "date_of_loss", "tgl_kejadian", "tanggal_kejadian", "loss_date", "tgl_klaim"],
            "source_direct_coins_fac": ["source_direct_coins_inward_fac", "source_direct_coins_fac", "source", "sumber_bisnis"],
            "curr": ["curr", "currency", "mata_uang", "valuta"],
            "claim_100": ["claim_100", "100_claim", "total_claim", "claim_100percent", "100_claim_amount"],
            "cedants_share_percent": ["cedants_share_percent", "cedants_share", "share_cedant_percent", "cedant_share_percent", "share_cedant_persen"],
            "cedants_share_in_amount": ["cedants_share_in_amount", "cedant_share_amount", "nilai_share_cedant", "nominal_cedant_share", "cedants_share_amount"],
            "spreading_of_claim_or": ["spreading_of_claim_or", "claim_or", "spreading_or"],
            "spreading_of_claim_qs": ["spreading_of_claim_qs", "claim_qs", "spreading_qs"],
            "spreading_of_claim_spl": ["spreading_of_claim_surplus", "spreading_of_claim_spl", "spreading_spl"],
            "spreading_of_claim_others": ["spreading_of_claim_others", "spreading_others", "claim_others"],
            "claim_qs_marsh_re_share": ["claim_quota_share_marsh_re_share", "claim_qs_marsh_re_share", "claim_qs_marsh"],
            "claim_spl_marsh_re_share": ["claim_surplus_marsh_re_share", "claim_spl_marsh_re_share", "claim_spl_marsh"],
            "claims_marein_share": ["claims_marein_share", "claim_marein_share", "marein_share"],
            "os_claims_100": ["outstanding_claims_100", "os_claims_100", "os_claim_100", "outstanding_claims_100percent"],
            "os_claims_marsh_re_share": ["outstanding_claims_marsh_re_share", "os_claims_marsh_re_share", "os_marsh"],
            "os_claims_marein_share": ["outstanding_claims_marein_share", "os_claims_marein_share", "os_marein"],
            "note": ["note", "notes", "keterangan", "catatan", "remark", "remarks"]
        }

        # 5. Smart Matcher: Exact Match Prioritas
        rename_dict = {}
        for orig_col in df.columns:
            s_col = to_snake_case(orig_col)
            matched_target = None

            for target_col, syn_list in SYNONYMS.items():
                if s_col == target_col or s_col in syn_list:
                    matched_target = target_col
                    break

            if not matched_target:
                for target_col, syn_list in SYNONYMS.items():
                    if any(syn in s_col for syn in syn_list):
                        matched_target = target_col
                        break

            if matched_target:
                rename_dict[orig_col] = matched_target

        df.rename(columns=rename_dict, inplace=True)
        df = df.loc[:, ~df.columns.duplicated()].copy()

        # 6. Filter Baris Bersih & Grand Total
        for id_col in ["register_no", "policy_number"]:
            if id_col in df.columns:
                df[id_col] = df[id_col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
                df[id_col] = df[id_col].replace(['nan', 'None', 'NaN', 'NaT', '', '<NA>'], np.nan)

        df = df.dropna(subset=["register_no", "policy_number"], how="all")

        trash_exact_pattern = r'^\s*(TOTAL|JUMLAH|GRAND\s*TOTAL|SUBTOTAL|REKAP|SUMMARY|0|NAN|NONE)\s*$'
        if "register_no" in df.columns:
            reg_str = df["register_no"].astype(str).str.upper().str.strip()
            df = df[~reg_str.str.match(trash_exact_pattern, na=False)]

        # 7. Bersihkan UW YEAR
        if "uw_year" in df.columns:
            def clean_uw_year(val):
                if pd.isna(val) or str(val).strip() in ['', 'None', 'nan', 'NaN']:
                    return None
                m = re.search(r'\d{4}', str(val))
                return m.group() if m else str(val).replace('.0', '').strip()
            df["uw_year"] = df["uw_year"].apply(clean_uw_year)

        # 8. Tanggal Parsing (Standardized ISO String)
        date_cols = ["dol", "period_of_insurance_start", "period_of_insurance_end"]
        for dcol in date_cols:
            if dcol in df.columns:
                def clean_date_str(v):
                    if pd.isna(v) or v is None:
                        return None
                    try:
                        dt = pd.to_datetime(v, errors='coerce', dayfirst=True)
                        if pd.isna(dt) or dt.year < 1900 or dt.year > 2500:
                            return None
                        return dt.strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        return None
                df[dcol] = df[dcol].apply(clean_date_str)

        # 9. Set Periode & Override COB
        df['period'] = periode_lengkap
        if override_cob and override_cob.strip() and override_cob.strip().lower() != "string":
            if 'cob' in df.columns:
                df['cob'] = override_cob.strip().upper()

        # 10. Sinkronisasi Kolom Master
        for col in master_cols:
            if col not in df.columns:
                df[col] = np.nan

        df_clean = df[master_cols].copy()

        # 11. Sanitasi Angka & Nominal
        num_cols = [
            "claim_100", "cedants_share_percent", "cedants_share_in_amount",
            "spreading_of_claim_or", "spreading_of_claim_qs", "spreading_of_claim_spl",
            "spreading_of_claim_others", "claim_qs_marsh_re_share", "claim_spl_marsh_re_share",
            "claims_marein_share", "os_claims_100", "os_claims_marsh_re_share", "os_claims_marein_share"
        ]
        for col in num_cols:
            if col in df_clean.columns:
                df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0.0)

        # 12. KONVERSI KOLOM 'no' KE INTEGER MURNI (1, 2, 3...)
        if "no" in df_clean.columns:
            clean_no = df_clean["no"].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            df_clean["no"] = pd.to_numeric(clean_no, errors='coerce').astype('Int64')

        return df_clean.reset_index(drop=True)